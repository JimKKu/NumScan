import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def find_matching_numbers(targets, input_file="a.xlsx", output_file="b.xlsx"):
    try:
        # 读取 a.xlsx 文件的所有数据
        df = pd.read_excel(input_file, header=None)  # 假设没有表头
        all_values = df.values.flatten()  # 展平所有数据
        matching_data = {}

        # 查找每个 target 的匹配项
        for target in targets:
            try:
                target_num = float(target)  # 支持数字字符串
                matches = [val for val in all_values if val == target_num]
                matching_data[target_num] = matches
            except ValueError:
                print(f"⚠️ 无效的目标参数: '{target}' (将被忽略)")

        # 终端输出结果
        print("\n匹配结果:")
        for target, matches in matching_data.items():
            print(f"🔍 目标 {target}: 找到 {len(matches)} 个匹配 → {matches}")

        # 生成 b.xlsx 文件
        wb = load_workbook(input_file)  # 以 a.xlsx 为模板
        ws = wb.active
        ws.title = "匹配结果"

        # 写入目标及匹配数据（新工作表）
        output_ws = wb.create_sheet("汇总")
        row_idx = 1

        # 设置高亮样式
        highlight_font = Font(bold=True)
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for target, matches in matching_data.items():
            # 写入目标（加粗 + 黄底高亮）
            output_ws.cell(row=row_idx, column=1, value=f"目标: {target}")
            output_ws.cell(row=row_idx, column=1).font = highlight_font
            output_ws.cell(row=row_idx, column=1).fill = highlight_fill
            row_idx += 1

            # 写入匹配数据
            for match in matches:
                output_ws.cell(row=row_idx, column=1, value=match)
                row_idx += 1

        wb.save(output_file)
        print(f"\n✅ 结果已保存至 {output_file}")

    except FileNotFoundError:
        print(f"❌ 错误: 文件 {input_file} 不存在！")
    except Exception as e:
        print(f"❌ 发生错误: {e}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python search_targets.py <target1> <target2> ...")
    else:
        targets = sys.argv[1:]
        find_matching_numbers(targets)
