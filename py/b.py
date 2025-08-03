import pandas as pd
import sys
from itertools import combinations
from openpyxl import Workbook
from openpyxl.styles import Font

def find_combinations(targets, input_file="a.xlsx", output_file="b.xlsx", max_r=3):
    try:
        # 读取 a.xlsx 所有数字（保留2位小数）
        df = pd.read_excel(input_file, header=None)
        numbers = [round(num, 2) for num in df.values.flatten() if not pd.isna(num)]
        
        # 控制台输出
        print(f"\n📊 所有金额: {', '.join(map(lambda x: f'{x:.2f}', numbers))}")
        print(f"🎯 目标值: {', '.join(map(str, targets))}\n")

        # 搜索每个目标的第一个匹配组合
        result = {}
        for target in targets:
            try:
                target_num = round(float(target), 2)
                found = None
                for r in range(2, max_r + 1):  # 仅检查2到max_r的组合
                    for combo in combinations(numbers, r):
                        if abs(sum(combo) - target_num) < 0.001:  # 浮点数精度处理
                            found = combo
                            break  # 找到第一个后立即停止
                    if found:
                        break
                result[target_num] = list(found) if found else ["无"]
            except ValueError:
                print(f"⚠️ 忽略无效目标: {target}")

        # 控制台输出结果
        print("匹配结果:")
        for target, nums in result.items():
            if nums == ["无"]:
                print(f"❌ {target:.2f}: 无匹配组合")
            else:
                print(f"✅ {target:.2f}: {' + '.join(map(lambda x: f'{x:.2f}', nums))} = {target:.2f}")

        # 生成 b.xlsx（目标值加粗，数字横向展开）
        wb = Workbook()
        ws = wb.active
        ws.title = "匹配结果"
        bold_font = Font(bold=True)

        for row_idx, (target, nums) in enumerate(result.items(), start=1):
            # 写入目标值（A列，加粗，不显示"Target"文字）
            ws.cell(row=row_idx, column=1, value=target).font = bold_font

            # 写入组合数字（B列、C列...，不加粗）
            if nums == ["无"]:
                ws.cell(row=row_idx, column=2, value="无")
            else:
                for col_idx, num in enumerate(nums, start=2):
                    ws.cell(row=row_idx, column=col_idx, value=num)

        wb.save(output_file)
        print(f"\n💾 结果已保存至 {output_file}")

    except FileNotFoundError:
        print(f"❌ 错误: {input_file} 不存在！")
    except Exception as e:
        print(f"❌ 发生错误: {e}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python find_combinations_final.py <目标值1> <目标值2> ... [--max_r 3]")
    else:
        # 解析命令行参数
        targets = []
        max_r = 10
        i = 1
        while i < len(sys.argv):
            if sys.argv[i] == "--max_r":
                max_r = int(sys.argv[i + 1])
                i += 2
            else:
                targets.append(sys.argv[i])
                i += 1
        find_combinations(targets, max_r=max_r)